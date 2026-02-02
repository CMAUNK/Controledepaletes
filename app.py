import streamlit as st
from openpyxl import load_workbook
from datetime import date
import pandas as pd
import re
import io

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm

# ================= CONFIG =================
st.set_page_config(page_title="Controle de Paletes", layout="centered")

BASE_FILE = "planilha_base.xlsx"
CODE_COLUMN = "B"
QTY_COLUMN = "C"
DATE_CELL = "C1"

# ================= FUNÇÕES =================

def carregar_codigos_base():
    wb = load_workbook(BASE_FILE)
    ws = wb.active
    codigos = []

    for row in range(3, ws.max_row + 1):
        code = ws[f"{CODE_COLUMN}{row}"].value
        if code:
            codigos.append(code)

    return codigos


def normalizar_texto(texto):
    pares = re.findall(
        r"(S\d{2})\s*(?:-|,)?\s*(\d+)",
        texto.upper()
    )

    dados = {}
    for codigo, quantidade in pares:
        dados[codigo] = int(quantidade)

    return dados


def gerar_planilha(dados, data_str):
    wb = load_workbook(BASE_FILE)
    ws = wb.active

    ws[DATE_CELL] = data_str

    for row in range(3, ws.max_row + 1):
        code = ws[f"{CODE_COLUMN}{row}"].value
        ws[f"{QTY_COLUMN}{row}"].value = dados.get(code, 0)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def gerar_pdf(df, data_str):
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )

    elementos = []

    # Cabeçalho
    header = Table(
        [["CONTROLE DE PALETES", f"DATA: {data_str}"]],
        colWidths=[12*cm, 4*cm]
    )

    header.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.grey),
        ("TEXTCOLOR", (0,0), (-1,-1), colors.white),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("FONT", (0,0), (-1,-1), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 14),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
    ]))

    elementos.append(header)

    # Dados da tabela
    dados_tabela = [df.columns.tolist()] + df.values.tolist()

    tabela = Table(
        dados_tabela,
        colWidths=[8*cm, 3*cm, 3*cm]
    )

    tabela.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 1, colors.black),
        ("BACKGROUND", (0,0), (-1,0), colors.red),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("ALIGN", (1,1), (-1,-1), "CENTER"),
        ("FONT", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONT", (0,1), (-1,-1), "Helvetica"),
    ]))

    elementos.append(tabela)
    doc.build(elementos)

    buffer.seek(0)
    return buffer

# ================= UI =================

st.title("📦 Controle de Paletes por Voz / Texto")
st.markdown("Modelo fixo | Pré-visualização obrigatória | Excel real")

texto = st.text_area(
    "Digite ou cole os códigos (ex: S21 - 6, S31 - 9)",
    height=150
)

data_sel = st.date_input("Data", value=date.today())
data_str = data_sel.strftime("%d/%m/%Y")

if st.button("Interpretar"):
    codigos_base = carregar_codigos_base()
    dados_brutos = normalizar_texto(texto)

    dados_validos = {
        c: q for c, q in dados_brutos.items()
        if c in codigos_base
    }

    if not dados_validos:
        st.error("Nenhum código válido encontrado.")
    else:
        st.session_state["dados"] = dados_validos
        st.session_state["confirmar"] = True

# ================= PRÉ-VISUALIZAÇÃO =================

if st.session_state.get("confirmar"):

    st.subheader("🧾 Pré-visualização (editável)")

    df_preview = pd.DataFrame(
        [{"Código": c, "Quantidade": q}
         for c, q in st.session_state["dados"].items()]
    )

    df_editado = st.data_editor(df_preview, num_rows="fixed")

    if st.button("Confirmar e gerar planilha"):

        dados_finais = dict(
            zip(df_editado["Código"], df_editado["Quantidade"])
        )

        arquivo_excel = gerar_planilha(dados_finais, data_str)

        st.success("Planilha gerada com sucesso")

        st.download_button(
            label="⬇️ Baixar planilha",
            data=arquivo_excel,
            file_name=f"CONTROLE_DE_PALETES_{data_str.replace('/', '-')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Visualização fiel ao modelo
        arquivo_excel.seek(0)
        df_visualizacao = pd.read_excel(
            arquivo_excel,
            header=1,
            usecols="A:C"
        ).reset_index(drop=True)

        # Remove linhas vazias (fix do nan nan 0)
        df_visualizacao = df_visualizacao[
            df_visualizacao["UNIDADE"].notna()
        ]

        # Normaliza cabeçalho
        df_visualizacao.columns = [
            "UNIDADE",
            "CÓDIGO",
            "QUANTIDADE DE PALETES"
        ]

        st.subheader("👀 Visualização da planilha final")
        st.dataframe(
            df_visualizacao,
            use_container_width=True,
            height=600
        )

        # PDF
        pdf = gerar_pdf(df_visualizacao, data_str)

        st.download_button(
            label="⬇️ Baixar PDF",
            data=pdf,
            file_name=f"CONTROLE_DE_PALETES_{data_str.replace('/', '-')}.pdf",
            mime="application/pdf"
        )
